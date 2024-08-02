#**************************************************************************************************************************************************************
# Proyecto de análisis
#       Contenido: Dataset NetflixPrize
#                  Dashboard Power BI
#                  Tratamiento de datos: Analisis.xlsx
#**************************************************************************************************************************************************************

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.comments import Comment

class AnalisisPeliculas:
    def __init__(self, ruta_datos):
        self.ruta_datos = ruta_datos
        self.wb = Workbook()
        self.pelicula_dict, self.year_pelicula_dict = self.cargarTitulos()
        self.crearHojaDetalle()
    
    def cargarTitulos(self):
        ruta_csv = os.path.join(self.ruta_datos, "data", "movie_titles.csv")
        print(f"Ruta completa del archivo CSV: {ruta_csv}")
        pelicula_dict = {}
        year_pelicula_dict = {}
        try:
            with open(ruta_csv, 'r', encoding='latin1') as archivo_csv:
                for linea in archivo_csv:
                    partes = linea.strip().split(',', 2)
                    if len(partes) == 3:
                        id_pelicula = int(partes[0])
                        year_pelicula = partes[1] if partes[1].lower() != 'null' else ""
                        titulo = partes[2]
                        
                        pelicula_dict[id_pelicula] = titulo
                        year_pelicula_dict[id_pelicula] = year_pelicula
        except FileNotFoundError:
            print(f"Error: No se encontró el archivo {ruta_csv}")
        except Exception as e:
            print(f"Error al leer el archivo {ruta_csv}: {e}")
        return pelicula_dict, year_pelicula_dict
    
    def crearHojaDetalle(self):
        if "Sheet" in self.wb.sheetnames:
            hoja_predeterminada = self.wb["Sheet"]
            self.wb.remove(hoja_predeterminada)
        
        self.detalle_ws = self.wb.create_sheet(title="Detalle")
        self.detalle_ws.append(["IdMovie", "MovieTitle", "PremiereYear", "C1", "C2", "C3", "C4", "C5", "TotalRatings", "StartDate", "EndDate", "Media", "Estacion"])
    
    def calcularPromedioFecha(self, fecha_inicio, fecha_fin):
        if fecha_inicio and fecha_fin:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
            fecha_promedio_dt = fecha_inicio_dt + (fecha_fin_dt - fecha_inicio_dt) / 2
            return fecha_promedio_dt.strftime("%d-%m-%Y")
        return ""
    
    def obtenerEstacion(self, fecha):        
        if fecha:
            fecha_dt = datetime.strptime(fecha, "%d-%m-%Y")
            
            primavera_inicio = datetime(fecha_dt.year, 9, 21)
            primavera_fin = datetime(fecha_dt.year, 12, 20)
            
            verano_inicio = datetime(fecha_dt.year, 12, 21)
            verano_fin = datetime(fecha_dt.year + 1, 3, 19)
            
            otoño_inicio = datetime(fecha_dt.year, 3, 20)
            otoño_fin = datetime(fecha_dt.year, 6, 20)
            
            invierno_inicio = datetime(fecha_dt.year, 6, 21)
            invierno_fin = datetime(fecha_dt.year, 9, 20)
            
            if primavera_inicio <= fecha_dt <= primavera_fin:
                return "Primavera"
            elif verano_inicio <= fecha_dt <= verano_fin:
                return "Verano"
            elif otoño_inicio <= fecha_dt <= otoño_fin:
                return "Otoño"
            elif invierno_inicio <= fecha_dt <= invierno_fin:
                return "Invierno"
            else:
                return "Invierno"
        return "Invierno"

    def procesarArchivo(self, ruta_archivo):
        id_pelicula = None
        conteo_calificaciones = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        fechas = []
        
        try:
            with open(ruta_archivo, "r") as archivo:
                for linea in archivo:
                    if linea.strip().endswith(":"):
                        if id_pelicula is not None:
                            self.escribirFilaDetalle(id_pelicula, conteo_calificaciones, fechas)
                        id_pelicula = int(linea.strip()[:-1])
                        conteo_calificaciones = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
                        fechas = []
                    else:
                        partes = linea.strip().split(',')
                        if len(partes) == 3:
                            calificacion = int(partes[1])
                            if calificacion in conteo_calificaciones:
                                conteo_calificaciones[calificacion] += 1
                            fechas.append(partes[2])
            if id_pelicula is not None:
                self.escribirFilaDetalle(id_pelicula, conteo_calificaciones, fechas)
        except FileNotFoundError:
            print(f"Error: No se encontró el archivo {ruta_archivo}")
        except Exception as e:
            print(f"Error al procesar el archivo {ruta_archivo}: {e}")
    
    def escribirFilaDetalle(self, id_pelicula, conteo_calificaciones, fechas):
        total_calificaciones = sum(conteo_calificaciones.values())
        titulo_pelicula = self.pelicula_dict.get(id_pelicula, "Unknown Title")
        year_estreno = self.year_pelicula_dict.get(id_pelicula, "")
        
        # Si el año de estreno es nulo, se asigna el año 1900 y se añade un comentario
        if not year_estreno:
            year_estreno = 1900
            comentario = Comment("Año de estreno nulo reemplazado por 1900 para fines de análisis", "Sistema")
        else:
            year_estreno = int(year_estreno)
            comentario = None

        fecha_inicio = min(fechas) if fechas else ""
        fecha_fin = max(fechas) if fechas else ""
        fecha_promedio = self.calcularPromedioFecha(fecha_inicio, fecha_fin)
        fecha_promedio_dt = datetime.strptime(fecha_promedio, "%d-%m-%Y") if fecha_promedio else None
        estacion = self.obtenerEstacion(fecha_promedio) if fecha_promedio else ""
        
        fila = [
            id_pelicula,
            titulo_pelicula,
            year_estreno,
            conteo_calificaciones[1],
            conteo_calificaciones[2],
            conteo_calificaciones[3],
            conteo_calificaciones[4],
            conteo_calificaciones[5],
            total_calificaciones,
            datetime.strptime(fecha_inicio, "%Y-%m-%d").strftime("%d-%m-%Y") if fecha_inicio else "",
            datetime.strptime(fecha_fin, "%Y-%m-%d").strftime("%d-%m-%Y") if fecha_fin else "",
            fecha_promedio,
            estacion
        ]
        
        self.detalle_ws.append(fila)
        
        # Si el año de estreno es nulo, se inserta el comentario
        if comentario:
            cell = self.detalle_ws.cell(row=self.detalle_ws.max_row, column=3)
            cell.comment = comentario

    
    def escribirDatosDetalle(self):
        nombres_archivos = ["combined_data_1.txt", "combined_data_2.txt", "combined_data_3.txt", "combined_data_4.txt"]
        for nombre_archivo in nombres_archivos:
            ruta_archivo = os.path.join(self.ruta_datos, "data", nombre_archivo)
            if os.path.exists(ruta_archivo):
                print(f"Procesando archivo: {ruta_archivo}")
                self.procesarArchivo(ruta_archivo)
    
    def guardarArchivo(self):
        ruta_salida = os.path.join(self.ruta_datos, "Informe/Analisis.xlsx")
        self.wb.save(ruta_salida)

if __name__ == "__main__":
    ubicacion = os.path.dirname(os.path.abspath(__file__))
    analisis = AnalisisPeliculas(ubicacion)
    analisis.escribirDatosDetalle()
    analisis.guardarArchivo()
